import pandas as pd
import psycopg2
import psycopg2.extras
import openpyxl

def connection(USERNAME, PASSWORD):    
    """
    Permet de se connecter au CREMI
    """

    print("Connexion à la base de données...")
    try:
        conn = psycopg2.connect(host="pgsql", dbname=USERNAME, user=USERNAME, password=PASSWORD)
    except Exception as e:
        exit("Unable to connect to the database: " + str(e))
        
    print("Connected to the database")
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    return conn, cur

def execute_query(command, conn, cur, print_output=False):
    """
    Execute une requete 
    """
    try:
        cur.execute(command)
        rows = cur.fetchall()
        if print_output:
            if rows:
                for row in rows:
                    print(row)
            else:
                print("Rien à afficher.")
        return rows
    except psycopg2.Error as e:
        cur.close()
        conn.close()
        print("Error:", e)
        return None
    
def display_regions(cur):
    cur.execute("SELECT id_reg, nom_reg FROM public.Regions")
    regions = cur.fetchall()
    print("Régions disponibles :")
    for i, region in enumerate(regions, 1):
        print(f"{region[0]}. {region[1]}")

def menu_reg(conn, cur): 
    print("\nChoisisez une région en donant son numéro :\n")
    display_regions(cur)   
    return int(input("Entrer le numéro de la région\n> "))
 
def display_departments_of_region(cur, chosen_region):
    cur.execute("SELECT id_dep, nom_dep FROM public.Departements WHERE id_reg = %s", (chosen_region,))
    departements = cur.fetchall()
    print("\nDépartements de la région :", chosen_region)
    for i, departement in enumerate(departements, 1):
        print(f"{departement[0]}. {departement[1]}")

def choix_niveau_etude():
    print("\nChoisisez le niveau de l'étude : \n")
    print("1: Régional (Economie, Social)")
    print("2: Départemental (Population)")
    while True : 
        try : 
            choix = int(input("> "))
            if choix == 1 :
                return "reg"
            if choix == 2 :
                return "dep"
        except :
            pass 
        print("Entrée non valide")

def choix_theme():
    print("\nChoisissez un thème :\n1. Population\n2. Social\n3. Economie\n")
    while True : 
        try : 
            choix = int(input("> "))
            if choix == 1 :
                return "Population"
            if choix == 2 :
                return "Social"
            if choix == 3 :
                return "Economie"
        except :
            pass 
        print("Entrée non valide")

def menu_dep(conn, cur, reg): 
    display_departments_of_region(cur, reg)   
    return input("\nEntrer le numéro de la région\n> ")

def choix_impression(cols, results): 
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    list_of_colnames = ["Zone"]+cols.split(',')
    if input("Ecrivez 1 pour sauvegarder la table\n> ") == '1':
        for col_idx, header in enumerate(list_of_colnames, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        for row_idx, row_data in enumerate(results, start=2):
            for col_idx, col_data in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=col_data)
        workbook.save("navigation_output.xlsx")

def choose_year(aviable_years):
    for i in aviable_years:
        print(i)
    return input("\nTapez une année\n> ")
    
def display_available_years(niveau_etude, theme):
    print("\nVoici les années disponnibles pour le theme selectionné :\n")
    if niveau_etude == "dep":
        if theme == "Population":
            return choose_year(["2015", "2020", "2023"])
    if niveau_etude == "reg":
        if theme == "Social":
            return choose_year(["2013", "2016", "2018", "2021"])
        if theme == "Economie":
            return choose_year(["2009", "2010", "2014", "2017", "2019"])
    print("\nAucune année pour le theme/niveau selectionné")
    return None 
    
def case_operator(year, theme):
    if theme == "Social":
        return {
            '2013': 'zone_inondable_2013',
            '2018': 'zone_inondable_2018',
            '2016': 's.egloignement_sante_2016',
            '2021': 's.eloignement_sante_2021',
        }.get(year, 'public.Regions')
    if theme == "Economie":
        return {
            '2009': 's.part_jeune_diplome_2009',
            '2019': 'taux_activite_2019',
            '2017': 'taux_activite_2017',
            '2014': 's.part_jeune_diplome_2014, effort_recherche_2014',
            '2010': 'effort_recherche_2010',
        }.get(year, 'public.Regions')
    return 'estimation_pop'

def display_data(zone, niveau_etude, theme, year, conn, cur): 
    if niveau_etude == "dep" :
        table = "estimation_pop_" + year
        query = "SELECT nom_dep, {} FROM public.Departements WHERE id_dep = %s".format(table)
        cur.execute(query, (zone,))
    
    if niveau_etude == "reg":
        table = case_operator(year, theme)
        if theme == "Social":
            query = """
                SELECT r.nom_reg, {}
                FROM public.Regions r
                JOIN public.Social s ON r.id_reg = s.id_reg
                WHERE r.id_reg = %s;
                """.format(table)

        if theme == "Economie":
            query = """
                SELECT r.nom_reg, {}
                FROM public.Regions r
                JOIN public.Economie s ON r.id_reg = s.id_reg
                WHERE r.id_reg = %s;
                """.format(table)
        cur.execute(query, (zone,))  
    results = cur.fetchall()
    for row in results:
        print(row)
    return results

""" Pour la connection en local """ 
conn = psycopg2.connect(
    host="localhost",
    dbname="public",
    user="postgres",
    password="superlucienpiat"
)
cur = conn.cursor()

""" Pour la connection au cremi """
# conn, cur = connection("username", "password")

### MAIN 

exit = ""
while exit!="q":
    zone = menu_reg(conn, cur)
    niveau_etude = choix_niveau_etude()
    if niveau_etude == 'dep':
        zone = menu_dep(conn, cur, zone)
    theme = choix_theme()
    year = display_available_years(niveau_etude, theme)
    if year :
        cols = case_operator(year, theme)
        print("Zone,", cols)
        results = display_data(zone, niveau_etude, theme, year, conn, cur)
        choix_impression(cols, results)
    exit = input("Appuyez sur q pour quiter: ")


