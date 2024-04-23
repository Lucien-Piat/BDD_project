import pandas as pd
import psycopg2 
import psycopg2.extras
import openpyxl

def connection(USERNAME, PASSWORD):    
    """
    Permet de se connecter au CREMI
    """

    print("Connexion à la base de données...")
    try:
        conn = psycopg2.connect(host="pgsql", dbname=USERNAME, user=USERNAME, password=PASSWORD)
    except Exception as e:
        exit("Unable to connect to the database: " + str(e))
        
    print("Connected to the database")
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    return conn, cur

def execute_query(command, conn, cur, print_output=False):
    """
    Execute une requete 
    """
    try:
        cur.execute(command)
        rows = cur.fetchall()
        if print_output:
            if rows:
                for row in rows:
                    print(row)
            else:
                print("Rien à afficher.")
        return rows
    except psycopg2.Error as e:
        cur.close()
        conn.close()
        print("Error:", e)
        return None
    
def output_to_excel(output_dict, file, list_of_colnames, numero_question):
    """
    Sauvegarde un dict sur un fichier excel
    """
    # On crée une page sur le fichier Excel 
    workbook = openpyxl.load_workbook(filename=file, data_only=True)
    sheet = workbook.create_sheet("Question_0" + str(numero_question))

    # Ajout des nom de collones
    for col_idx, header in enumerate(list_of_colnames, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Ajout des lignes
    for row_idx, row_data in enumerate(output_dict, start=2):
        for col_idx, col_data in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=col_data)

    workbook.save(file)

""" Pour la connection en local """ 
conn = psycopg2.connect(
    host="localhost",
    dbname="public",
    user="postgres",
    password="superlucienpiat"
)
cur = conn.cursor()

""" Pour la connection au cremi """
# conn, cur = connection("username", "password")

file = "reponses.xlsx"
workbook = openpyxl.Workbook()
workbook.save(file)

# Afficher la liste des régions où le taux de la population habitant en zone inondable 
# était supérieure a 10% en 2013, classées du plus fort taux au plus faible.

question_1 = """SELECT r.nom_reg, s.zone_inondable_2013
    FROM public.Regions r 
    JOIN public.Social s ON r.id_reg = s.id_reg 
    WHERE s.zone_inondable_2013 > 10 
    ORDER BY s.zone_inondable_2013 DESC;"""

output_dict = execute_query(question_1, conn, cur)
output_to_excel(output_dict, file, ["Region", 
                                    "taux de la population habitant en zone inondable en 2013"], 1)

# Quelle est la variation de l'effort de recherche et développement entre 2010 et 2014 
# dans les régions dont moins de 70% de la population a une utilisation quotidienne d'internet en 2019? 
# Afficher aussi le taux d'activité en 2019 pour ces régions.

question_2 = """SELECT r.nom_reg, (e.effort_recherche_2014 - e.effort_recherche_2010) AS variation_effort_recherche, 
    e.taux_activite_2019
    FROM public.Regions r 
    JOIN public.CompNum c ON r.id_reg = c.id_reg
    JOIN public.Economie e ON r.id_reg = e.id_reg
    WHERE c.cn_quotidienne < 70
    ORDER BY variation_effort_recherche;;"""

output_dict = execute_query(question_2, conn, cur)
output_to_excel(output_dict, file, ["Region",
                                    "Variation de l'effort de recherche et développement entre 2010 et 2014",
                                    "Taux d'activité en 2019"], 2)

# Quels sont les départements dont moins du quart la population de la région avait un des compétences numériques fortes en 2019 ? 
# Afficher aussi l'évolution de la part des jeunes diplôme entre 2009 et 2014 pour ces départements.

question_3 = """ SELECT d.nom_dep, d.id_dep,
    (e.part_jeune_diplome_2014 - e.part_jeune_diplome_2009) AS evolution_jeunes_diplomes 
    FROM public.Departements d 
    JOIN public.CompNum c ON d.id_reg = c.id_reg 
    JOIN public.Economie e ON d.id_reg = e.id_reg
    WHERE c.cn_forte < 25
    ORDER BY evolution_jeunes_diplomes;
    """

output_dict = execute_query(question_3, conn, cur)
output_to_excel(output_dict, file, ["Département",
                                    "cnn", 
                                    'Evolution de la part des jeunes diplôme entre 2009 et 2014'
                                    ], 3)

# Quelle est l'estimation de population en 2020 de tous les départements où l'absence d'équipement internet 
# est supérieure ou égale a 12.5% (en 2019).

Question_4 = "SELECT nom_dep, estimation_pop_2020 FROM Departements INNER JOIN Population ON Departements.id_dep = Population.id_dep INNER JOIN CompetencesNumeriques ON CompetencesNumeriques.id_reg = Departements.id_reg WHERE intensite = 'sans' and taux >= 12.5;"

question_4 = """SELECT d.nom_dep, d.estimation_pop_2020 
    FROM public.Departements d 
    JOIN public.CompNum c ON d.id_reg = c.id_reg 
    WHERE c.cn_forte < 25;
    """

output_dict = execute_query(question_4, conn, cur)
output_to_excel(output_dict, file, ["Département",
                                    "Estimation de population en 2020"], 4)

# Quelle était la part de la population ayant une utilisation quotidienne d'internet dans les régions où le taux d'activité 
# était >75% en 2019 et où la Part de la population éloignée de plus de 7 mn des services de santé de proximité était de moins de 5% en 2021.

question_5 = """SELECT r.nom_reg, c.cn_quotidienne 
    FROM public.Regions r 
    JOIN public.Social s ON r.id_reg = s.id_reg 
    JOIN public.CompNum c ON r.id_reg = c.id_reg 
    JOIN public.Economie e ON r.id_reg = e.id_reg 
    WHERE e.taux_activite_2019 > 75
        AND s.egloignement_sante_2021 < 5;
    """

output_dict = execute_query(question_5, conn, cur)
output_to_excel(output_dict, file, ["Région",
                                    "Part de la population éloignée de plus de 7 mn des services de santé en 2021"], 5)

workbook = openpyxl.load_workbook(filename=file, data_only=True)
first_sheet = workbook.worksheets[0]  
workbook.remove(first_sheet)
workbook.save(file)